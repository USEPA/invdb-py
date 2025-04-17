import chalicelib.src.AWS.S3.methods as s3_methods
import chalicelib.src.qc_analytics.jobs.download_recalculations_excel.queries as recalc_download_queries
import chalicelib.src.qc_analytics.jobs.recalculations_report.methods as recalc_report_methods
from openpyxl.styles import Font
from flask import make_response
from openpyxl import Workbook
from io import BytesIO
import re

def get_column_order(data_set_row_keys, type: "input" or "raw"): 
    preferred_key_column_order = ["sector", "subsector", "category", "sub_category_1", "sub_category_2", 
                                  "sub_category_3", "sub_category_4", "sub_category_5", "carbon_pool", "fuel1", "fuel2", 
                                  "geo_ref", "exclude", "crt_code", "id", "cbi_activity", "units", "ghg", "ghg_category", "gwp"]
    if type == "input": 
        preferred_key_column_order = ["data_key"] + preferred_key_column_order
    if type == "raw": 
        preferred_key_column_order = ["emissions_key", "raw_data_key"] + preferred_key_column_order
    year_column_pattern = re.compile(r"Y[0-9]{4}")
    columns = [column_name for column_name in preferred_key_column_order if column_name in data_set_row_keys]
    columns += [column_name for column_name in data_set_row_keys if not column_name in preferred_key_column_order and not year_column_pattern.match(column_name)]
    columns += sorted([column_name for column_name in data_set_row_keys if year_column_pattern.match(column_name)])
    return columns

def populate_qca_event_metadata_tab_data(qca_event_metadata: dict, sheet):
    row = 0
    column = 1

    def recursive_populate_object(content: dict):
        nonlocal row, column
        for key, value in content.items():
            row += 1
            sheet.cell(row=row, column=column, value=f"{key}:")
            if isinstance(value, dict):
                column += 1
                recursive_populate_object(value)
                column -= 1
            elif isinstance(value, list):
                column += 1
                recursive_populate_list(value)
                column -= 1
            else:
                sheet.cell(row=row, column=column + 1, value=value)

    def recursive_populate_list(content: list):
        nonlocal row, column
        for item in content:
            row += 1
            if isinstance(item, dict):
                recursive_populate_object(item)
            elif isinstance(item, list):
                recursive_populate_list(item)
            else:
                sheet.cell(row=row, column=column, value=item)

    recursive_populate_object(qca_event_metadata)

def populate_dataset_tab_data(dataset_data: list, qca_event_metadata, sheet):
    if len(dataset_data) == 0: 
        return

    # get the column order
    columns = get_column_order(dataset_data[0].keys(), "input")
    
    # write the header row (in bold)
    for column_num, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_num, value=column_name)
        cell.font = Font(bold=True)

    # Write the data rows
    for row_num, row_data in enumerate(dataset_data, start=2): 
        for column_num, column_name in enumerate(columns, start=1): 
            sheet.cell(row=row_num, column=column_num, value=row_data[column_name])


def populate_aggregate_results_tab_data(aggregate_results, qca_event_metadata, sheet):
    #columns: is_aggregate, group_key, ...data_key_columns, ...output_years
    max_level = len(qca_event_metadata["columns"])
    key_columns = [qca_event_metadata["ghgOption"]] + qca_event_metadata["columns"]
    year_columns = [key for key in aggregate_results[0]["children"][1]["data"].keys() if "recalc" in key]

    # write the header row (in bold)
    for column_num, column_name in enumerate(["is_aggregrate", "group_key"] + key_columns + year_columns, start=1):
        cell = sheet.cell(row=1, column=column_num, value=column_name)
        cell.font = Font(bold=True)

    row_num = 2
    def recurse_populate_object(object_data: dict, level: int): 
        nonlocal max_level, row_num
        sheet.cell(row=row_num, column=1, value=(level < max_level)) # is_aggregate
        sheet.cell(row=row_num, column=2, value=object_data["key"]) # group_key
        for col_num, key_column_value in enumerate(object_data["key"].split("/"), start=3): # all the key columns
            sheet.cell(row=row_num, column=col_num, value=key_column_value)
        for col_num, year_column_value in enumerate([("null" if value is None else value) for key, value in object_data["data"].items() if "recalc" in key], start=3+max_level): # all the year columns
            sheet.cell(row=row_num, column=col_num, value=year_column_value)
        row_num += 1
        if "children" in object_data:
            for child in object_data["children"]:
                recurse_populate_object(child, level + 1)

    for aggregate_group in aggregate_results:
        recurse_populate_object(aggregate_group, 0)


def populate_raw_results_tab_data(raw_results, qca_event_metadata, sheet):
    if len(raw_results) == 0:
        return 

    # get the column order
    columns = get_column_order(raw_results[0]["data"][0].keys(), "raw")

    # write the header row (in bold)
    for column_num, column_name in enumerate(["group_key"] + columns, start=1):
        cell = sheet.cell(row=1, column=column_num, value=column_name)
        cell.font = Font(bold=True)
    
    row_num = 2
    for raw_data_group in raw_results: 
        for data_row in raw_data_group["data"]: 
            sheet.cell(row=row_num, column=1, value=raw_data_group["key"])
            for column_num, column_name in enumerate(columns, start=2):
                sheet.cell(row=row_num, column=column_num, value=data_row[column_name])
            row_num += 1


def handle_recalculations_excel_download_request(qca_event_directory_handle: str, debug: bool, user_id: int, test_case_name: str=None, is_local_test: bool=False):
    # fetch the datasets from S3
    if debug is None: 
        debug = False

    folder_name = qca_event_directory_handle[qca_event_directory_handle.rfind('/') + 1:]
    s3_session = s3_methods.get_global_s3_session()

    # fetch all the files within the qca folder
    (qca_event_metadata,
    baseline_dataset_data,
    comparator_dataset_data, 
    aggregate_results,
    raw_results) = recalc_download_queries.fetch_qca_object_files_from_s3(s3_session, qca_event_directory_handle)

    if debug: # get the intermediate recalc dataframes (to put in their own tabs) when debug is true.
        baseline_recalc_dataframe, comparator_recalc_dataframe, _ = recalc_report_methods.get_recalc_dataframes(qca_event_directory_handle, s3_session)
        baseline_aggregate_dataframe, baseline_raw_dataframe = baseline_recalc_dataframe.aggregate_data, baseline_recalc_dataframe.raw_data
        comparator_aggregate_dataframe, comparator_raw_dataframe = comparator_recalc_dataframe.aggregate_data, comparator_recalc_dataframe.raw_data

    # split "sub_category_fuel_1" column into "sub_category_1" and "fuel1" columns
    sub_category_fuel_1_index = qca_event_metadata['columns'].index('sub_category_fuel_1')
    qca_event_metadata['columns'] = qca_event_metadata['columns'][:sub_category_fuel_1_index] + ['sub_category_1', 'fuel1'] + qca_event_metadata['columns'][sub_category_fuel_1_index + 1:]

    # create the excel file
    workbook = Workbook()
    first_sheet_created = False

    if not debug: 
        tab_names_and_data = { 
            "Metadata": qca_event_metadata,
            "Baseline Input": baseline_dataset_data,
            "Comparator Input": comparator_dataset_data,
            "Aggregate Results": aggregate_results,
            "Raw Results": raw_results
        }
    else: 
        tab_names_and_data = { 
            "Metadata": qca_event_metadata,
            "Baseline Input": baseline_dataset_data,
            "Comparator Input": comparator_dataset_data,
            "Baseline Aggregate Dataframe": baseline_aggregate_dataframe,
            "Comparator Aggregate Dataframe": comparator_aggregate_dataframe,
            "Aggregate Results": aggregate_results,
            "Baseline Raw Dataframe": baseline_raw_dataframe,
            "Comparator Raw Dataframe": comparator_raw_dataframe,
            "Raw Results": raw_results
        }

    # create and populate the tabs for those that have data
    for tab_name, tab_data in tab_names_and_data.items():
        # create the new tab sheet and open it
        if tab_data is None: 
            continue

        if not first_sheet_created: 
            sheet = workbook.active 
            sheet.title = tab_name
            first_sheet_created = True
        else: 
            sheet = workbook.create_sheet(title=tab_name)
            workbook.active = sheet

        # call the appropriate function to populate that tab's data
        if tab_name == "Metadata":
            populate_qca_event_metadata_tab_data(qca_event_metadata, sheet)
        elif tab_name == "Baseline Input":
            populate_dataset_tab_data(baseline_dataset_data, qca_event_metadata, sheet)
        elif tab_name == "Comparator Input":
            populate_dataset_tab_data(comparator_dataset_data, qca_event_metadata, sheet)
        elif tab_name == "Aggregate Results":
            populate_aggregate_results_tab_data(aggregate_results, qca_event_metadata, sheet)
        elif tab_name == "Baseline Aggregate Dataframe":
            populate_aggregate_results_tab_data(baseline_aggregate_dataframe, qca_event_metadata, sheet)
        elif tab_name == "Comparator Aggregate Dataframe":
            populate_aggregate_results_tab_data(comparator_aggregate_dataframe, qca_event_metadata, sheet)
        elif tab_name == "Baseline Raw Dataframe":
            populate_raw_results_tab_data(baseline_raw_dataframe, qca_event_metadata, sheet)
        elif tab_name == "Comparator Raw Dataframe":
            populate_raw_results_tab_data(comparator_raw_dataframe, qca_event_metadata, sheet)  
        elif tab_name == "Raw Results":
            populate_raw_results_tab_data(raw_results, qca_event_metadata, sheet)
       
              

    # save and transmit the excel file
    file_bytes = BytesIO()
    workbook.save(file_bytes)
    file_bytes.seek(0)

    response = make_response(file_bytes.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f'attachment; filename="QC Analytics Object.xlsx"'

    return response
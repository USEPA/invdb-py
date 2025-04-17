import copy
import io
import os
import tempfile
from openpyxl.reader.excel import load_workbook
from chalicelib.src.general.helpers import full_class_name, tprint
import chalicelib.src.reports.constants as report_constants

class ReportRow:
    def __init__(self, sheet, row_number, width, read_only):
        self._sheet = sheet
        self.row_number = row_number
        self._width = width
        self._is_read_only = read_only

    def __getitem__(self, slice_obj):
        if isinstance(slice_obj, slice):
            start = slice_obj.start
            stop = slice_obj.stop
            step = slice_obj.step
            if step is None:
                step = 1
            if start is None:
                start = 1
            if stop is None:
                stop = self._width + 1

            return_value = [
                self._sheet.cell(
                    row=self.row_number, column=column + 1
                ).value  # cell() uses 1-based indexes
                for column in range(start, stop, step)
            ]
            if self._is_read_only:
                return copy.deepcopy(return_value)
            else:
                return return_value
        else:  # if only retrieving one element
            column = slice_obj + 1  # cell() uses 1-based indexes
            return_value = self._sheet.cell(row=self.row_number, column=column).value
            if self._is_read_only:
                return copy.deepcopy(return_value)
            else:
                return return_value

    def __setitem__(self, slice_obj, value):
        if self._is_read_only:
            raise PermissionError(
                f"{full_class_name(self)}.__setitem__: cannot mutate cells of a read only report object."
            )
        if isinstance(slice_obj, slice):
            start = slice_obj.start
            stop = slice_obj.stop
            step = slice_obj.step
            if step is None:
                step = 1
            if start is None:
                start = 1
            if stop is None:
                stop = self._width + 1
            for col, val in zip(range(start, stop, step), value):
                self._sheet.cell(row=self.row_number, column=col).value = val
        else:
            column = slice_obj + 1
            self._sheet.cell(row=self.row_number, column=column).value = value

    def __iter__(self):
        if self._is_read_only:
            yield from [
                copy.deepcopy(cell.value)
                for cell in self._sheet[self.row_number][: self._width + 1]
            ]
        else:
            yield from [
                cell.value
                for cell in self._sheet[self.row_number][: self._width + 1]
            ]

    def __len__(self):
        return self._width

    def __repr__(self):
        return f"<ReportRow Object: {[cell.value for cell in self._sheet[self.row_number][: self._width + 1]]}>"


class Report:
    def __init__(
        self,
        report_id,
        content,
        max_time_series,
        created_by=None,
        read_only=False,
        report_name=None,
    ):
        self._report_id = report_id
        self._content = content
        self._max_time_series = max_time_series
        self._created_by = created_by
        self._is_read_only = read_only
        self._report_name = report_name
        
        self._width = None
        self._workbook = None
        self._sheet = None
        self._temp_file = None
        self._is_open = False
        self._report_type = None
        self._length = len(self) # compute the length of the file


    def get_report_name(self):
        return self._report_name


    def get_report_id(self):
        return self._report_id


    def get_report_type(self):
        if self._report_type is None:
            return report_constants.REPORT_TYPES["UNSPECIFIED"]
        else:
            return self._report_type


    def get_max_time_series(self):
        return self._max_time_series


    def get_created_by(self):
        return self._created_by


    def is_read_only(self):
        return self._is_read_only


    def delete_rows(self, first_row, row_count):
        if not self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.delete_rows: a report must be opened before its contents can be accessed."
            )
        self._sheet.delete_rows(first_row, row_count)

    def insert_rows(self, first_row, row_count):
        if not self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.insert_rows: a report must be opened before its contents can be accessed."
            )
        self._sheet.insert_rows(first_row, row_count)

    def open(self):
        if self._is_open:
            tprint(
                f"Warning: There was an attempt to open a report that is already open: {self}"
            )
            return
        else:
            self.__enter__()

    def has_query_results_tab(self):
        return report_constants.REPORT_OUTPUT_DATA_SHEET_NAME in self._workbook.sheetnames

    def switch_to_queries_tab(self):
        # switch the 'Queries' tab
        self._workbook.active = self._workbook[report_constants.REPORT_INPUT_DATA_SHEET_NAME]
        self._sheet = self._workbook.get_sheet_by_name(report_constants.REPORT_INPUT_DATA_SHEET_NAME)

    def switch_to_results_tab(self):
        # create the 'Query_Results' tab if it does not already exist. Then switch to it.
        if not self.has_query_results_tab():
            self._workbook.create_sheet(title=report_constants.REPORT_OUTPUT_DATA_SHEET_NAME)
        self._workbook.active = self._workbook[report_constants.REPORT_OUTPUT_DATA_SHEET_NAME]
        self._sheet = self._workbook[report_constants.REPORT_OUTPUT_DATA_SHEET_NAME]

    def save(self):
        if not self._is_open:
            tprint(f"Warning: There was an attempt to save a closed report: {self}")
        self._workbook.save(self._temp_file.name)

    def save_to(self, target):
        if not self._is_open:
            tprint(f"Warning: There was an attempt to save a closed report: {self}")
        self._workbook.save(target)

    def close(self):
        if not self._is_open:
            tprint(
                f"Warning: There was an attempt to close a report that is already closed: {self}"
            )
            return
        else:
            self.__exit__("", "", "")

    def __len__(self):
        """return the number of queries in the input tab of the report. Includes invalid lines and non-trailing blank lines"""
        if hasattr(self, "_length"):
            return self._length
        else:
            with self as opened_self:
                formulas = self._sheet[report_constants.FORMULA_COL_LABEL][1:] # get the formula rows as an array (excluding the header row)
        
                last_formula = 0
                for formula, i in zip(formulas, range(1,len(formulas)+1)):
                    if formula.value and not formula.value.isspace():
                        last_formula = i

            return last_formula
            

    def __iter__(self):
        if not self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.__iter__: a report must be opened before it can be iterated."
            )
        
        row_num = 1
        for row in self._sheet.iter_rows(values_only=True):

            #skip the first row containing the column headers
            if row_num < report_constants.FIRST_DATA_ROW:
                row_num += 1
                continue

            # stop after the last row of true content
            if row_num > self._length + 1:
                return
            
            # return the next row
            yield ReportRow(self._sheet, row_num, self._width, self._is_read_only)
            row_num += 1

    def __getitem__(self, row):
        """retrieves a row from the sheet (note: the row indexes are 1-based)"""
        if not self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.__getitem__: a report must be opened before its contents can be accessed."
            )
        if not isinstance(row, int):
            raise TypeError(
                f"{full_class_name(self)}.__getitem__: a report indexes must be integers"
            )
        return ReportRow(self._sheet, row, self._width, self._is_read_only)

    def __setitem__(self, _):
        """retrieves a row from the sheet (note: the row indexes are 1-based)"""
        raise AttributeError(
            f"{full_class_name(self)}.__setitem__: row-wise mutations in reports are not supported. Use a second index to mutate slices or individual cells of the row."
        )

    def __enter__(self):
        if self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.__enter__: this file is already open."
            )
        try:
            if self._temp_file is None: # create the temp file if it does not exist yet
                self._temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                self._temp_file.write(io.BytesIO(self._content).getbuffer())
                self._temp_file.close()

            self._workbook = load_workbook(self._temp_file.name)
            self._sheet = self._workbook[report_constants.REPORT_INPUT_DATA_SHEET_NAME]
            self._is_open = True
        except Exception as e:
            tprint(f"Failed to open workbook: {e}")
            raise

        return self

    def __exit__(self, exc_type, exc_value, traceback):
        try:
            if self._workbook is not None:
                self._workbook.save(self._temp_file.name)
                self._workbook.close()
                self._is_open = False
        except Exception as e:
            tprint(f"Failed to save/close workbook: {e}")
            raise

        self._workbook = None
        self._sheet = None

        # Remove the temporary file
        # if self._temp_file is not None and os.path.exists(self._temp_file.name):
        #     os.remove(self._temp_file.name)

    def __repr__(self):
        return f"""<Report Object: report ID: {self._report_id}, report name: {self._report_name}>"""

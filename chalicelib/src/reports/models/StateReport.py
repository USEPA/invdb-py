from chalicelib.src.reports.models.Report import *
from chalicelib.src.jobs.models.Job import *
import chalicelib.src.reports.jobs.processing.queries as processing_queries
from chalicelib.src.general.helpers import full_class_name, tprint
import chalicelib.src.reports.constants as report_constants
from openpyxl.reader.excel import load_workbook
import math


class StateReportRow(ReportRow):
    pass

class StateReport(Report):
    def __init__(
        self,
        report_id,
        content,
        max_time_series,
        created_by=None,
        read_only=False,
        report_name=None,
    ):
        super().__init__(
            report_id,
            content,
            max_time_series,
            created_by,
            read_only,
            report_name,
        )
        self._report_type = report_constants.REPORT_TYPES["STATE"]
        self._width = report_constants.STATE_Y1990_QUANTITY_COL_POS + 1 + self._max_time_series
    

    def __repr__(self):
        return "<State " + super().__repr__()[1:]


    def get_input_queries_info(self) -> [list]:
        """returns a 1-based index list of the content of columns A-C in the Queries tab. This holds all the original query information (even ones with errors)"""
        if not self._is_open:
            raise ValueError(f"{helpers.full_class_name(self)}.get_queries_input(): Report must be open to get the queries input")

        active_worksheet_save = self._workbook.active.title
        self.switch_to_queries_tab()
        
        input_queries_info = self._sheet.iter_rows(
            min_row=1, 
            max_row=len(self)+1, 
            min_col=report_constants.STATE_AGGREGATESTO_COL_POS, 
            max_col=report_constants.STATE_COL_POS
        )
        
        self._sheet = self._workbook[active_worksheet_save]

        # format the return object as a 2D list of lists
        return_array = [''] # first row is a placeholder to make indexes 1-based
        for row in input_queries_info:
            current_row = []
            for cell in row:
                current_row.append(cell.value)
            return_array.append(current_row)
            
        return return_array


    def process_contents(self, error_rows: [int], query_formula_info, reporting_year: int, layer_id: int, this_job: Job) -> None:
        """takes a single national report and the list of its validation error rows, and updates the contents
        of the state report file to include the processed query results in the Query_Results tab"""
        self.switch_to_queries_tab()

        # wipe the existing data in the Query_Results tab
        if self.has_query_results_tab():
            self._workbook.remove(self._workbook[report_constants.REPORT_OUTPUT_DATA_SHEET_NAME])
        self._workbook.create_sheet(title=report_constants.REPORT_OUTPUT_DATA_SHEET_NAME)

        # gather the input query info
        all_queries_info = self.get_input_queries_info()
        # gather the object of all the valid queries and input column values
        valid_queries_info: {int: tuple} = {} # structure: {row number: (parsed formula info)}
        for row in [row for row, i in zip(self, range(report_constants.FIRST_DATA_ROW,len(self)+2)) if i not in error_rows]:
            formula_string = row[report_constants.FORMULA_COL_POS]
            if formula_string is not None:
                formula_string = formula_string.strip()
            if formula_string in ["", None]:
                continue # skip empty rows if not marked as error_row
            formula_prefix = formula_string[:formula_string.index("(")]
            input_arguments, _ = helpers.parse_report_formula_arguments(formula_string)
            valid_queries_info.update({row.row_number: (formula_prefix,) + tuple(input_arguments)})
        
        
        self.switch_to_results_tab()
        output_row_num = report_constants.FIRST_DATA_ROW
        resultless_query_rows =  []
        
        # process the queries in batches (batch size based on report_constants.QUERIES_PER_REQUEST)
        batch_number = 1
        batch_size = report_constants.QUERIES_PER_REQUEST
        batch_count = math.ceil((len(valid_queries_info.keys())) / batch_size)
        query_rows = sorted(valid_queries_info.keys())
        query_count = len(query_rows)
        for batch_head in range(0, query_count, batch_size): # batch head is the index for the list of valid queries
            this_job.post_event(
                "REPORT_PROCESSING",
                "PROCESSING_QUERY_BATCH",
                batch_number, 
                batch_count,
                self._report_id
            )
            # gather the batch of queries
            if batch_head + report_constants.QUERIES_PER_REQUEST < query_count: # full batch
                query_batch = [(query_rows[i],) + valid_queries_info[query_rows[i]] for i in range(batch_head, batch_head + report_constants.QUERIES_PER_REQUEST)]
            else: # final, partial batch
                query_batch = [(query_rows[i],) + valid_queries_info[query_rows[i]] for i in range(batch_head, query_count)]
            
            input_rows_this_batch = [query[0] for query in query_batch]

            if len(query_batch) == 0: # if the this upcoming batch is empty, stop 
                break

            # send the query batch to the database in one SQL statement and get results back
            # query_batch_results returns table with record structure: (row: int, year: int, state: str, total_quantity: Decimal)
            query_batch_results, width = processing_queries.process_report_query_batch(query_batch, query_formula_info, self._report_type, reporting_year, layer_id)
            
            # restructure the results to simplify formating output data row by row
            rows_with_results = []
            results_by_row = {}
            for row, year, state, total_quantity in query_batch_results:
                key = (row, state)
                rows_with_results.append(row)
                rows_with_results 
                if key not in results_by_row:
                    results_by_row[key] = []
                results_by_row[key].append((year, total_quantity))

            resultless_query_rows += [row for row in input_rows_this_batch if row not in rows_with_results]

            if batch_number == 1: # first batch only
                # write in the Query_Results tab column headers
                self[1][report_constants.STATE_AGGREGATESTO_COL_POS] = report_constants.STATE_REPORT_AGGREGATESTO_COLUMN_HEADER
                self[1][report_constants.STATE_AGGREGATIONSIGNANDFACTOR_COL_POS] = report_constants.STATE_REPORT_AGGREGATIONSIGNANDFACTOR_COLUMN_HEADER
                self[1][report_constants.FORMULA_COL_POS] = report_constants.REPORT_FORMULA_COLUMN_HEADER
                self[1][report_constants.STATE_ROW_IN_NUMBER_POS] = report_constants.STATE_REPORT_ROW_IN_NUMBER_COLUMN_HEADER
                self[1][report_constants.STATE_COL_POS] = report_constants.STATE_REPORT_STATE_COLUMN_HEADER
                for year, i in zip(range(db_constants.EARLIEST_REPORTING_YEAR,db_constants.EARLIEST_REPORTING_YEAR + width), range(1, width+1)):
                    self[1][report_constants.STATE_COL_POS + i] = f'Y{year}'

                # write in one line for the invalid queries (no state or quantity values)
                for errored_query_row_num in error_rows:
                    self[output_row_num][report_constants.STATE_AGGREGATESTO_COL_POS] = all_queries_info[errored_query_row_num][0]
                    self[output_row_num][report_constants.STATE_AGGREGATIONSIGNANDFACTOR_COL_POS] = all_queries_info[errored_query_row_num][1]
                    self[output_row_num][report_constants.FORMULA_COL_POS] = all_queries_info[errored_query_row_num][2]
                    self[output_row_num][report_constants.STATE_ROW_IN_NUMBER_POS] = errored_query_row_num
                    output_row_num += 1

            # write the results into their corresponding rows in the Query_Results tab
            this_job.post_event(
                "REPORT_PROCESSING",
                "WRITING_BATCH_RESULTS",
                batch_number, 
                batch_count,
                self._report_id
            )

            for key, result_series in results_by_row.items():
                input_row_num, state = key
                self[output_row_num][report_constants.STATE_AGGREGATESTO_COL_POS] = all_queries_info[input_row_num][0]
                self[output_row_num][report_constants.STATE_AGGREGATIONSIGNANDFACTOR_COL_POS] = all_queries_info[input_row_num][1]
                self[output_row_num][report_constants.FORMULA_COL_POS] = all_queries_info[input_row_num][2]
                self[output_row_num][report_constants.STATE_ROW_IN_NUMBER_POS] = input_row_num
                self[output_row_num][report_constants.STATE_COL_POS] = state
                for year, result in result_series:
                    self[output_row_num][report_constants.STATE_Y1990_QUANTITY_COL_POS + (year - db_constants.EARLIEST_REPORTING_YEAR)] = float(result if result else 0) * float(all_queries_info[input_row_num][1] if all_queries_info[input_row_num][1] else 1)
                output_row_num += 1

            batch_number += 1
            # END OF BATCH PROCESSING

        # underneath the error query lines, write in one line for the resultless queries (no state value and all 0 quantities)
        if len(resultless_query_rows) > 0:
            self.insert_rows(report_constants.FIRST_DATA_ROW + len(error_rows), len(resultless_query_rows))
            output_row_num = report_constants.FIRST_DATA_ROW + len(error_rows)
            for resultless_query_row_num in resultless_query_rows:
                self[output_row_num][report_constants.STATE_AGGREGATESTO_COL_POS] = all_queries_info[resultless_query_row_num][0]
                self[output_row_num][report_constants.STATE_AGGREGATIONSIGNANDFACTOR_COL_POS] = all_queries_info[resultless_query_row_num][1]
                self[output_row_num][report_constants.FORMULA_COL_POS] = all_queries_info[resultless_query_row_num][2]
                self[output_row_num][report_constants.STATE_ROW_IN_NUMBER_POS] = resultless_query_row_num
                for year_column in range(report_constants.STATE_Y1990_QUANTITY_COL_POS, report_constants.STATE_Y1990_QUANTITY_COL_POS + width):
                    self[output_row_num][year_column] = 0
                output_row_num += 1

        helpers.tprint("Saving the report...")
        self.save()
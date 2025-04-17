from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from chalicelib.src.general.helpers import full_class_name
import chalicelib.src.source_files.constants as source_file_constants
import chalicelib.src.general.qc_constants as qc_constants
import chalicelib.src.general.helpers as helpers 
import copy
import io
import os
import tempfile
import csv
import json

def strip_whitespace(value):
    """strip all trailing and leading whitespace from read string values. 
    Any whitespace-only values return 'NULL' """
    if isinstance(value, str): 
        return "NULL" if value.strip() == "" else value.strip()
    else:
        return value

class SourceFileRow:
    def __init__(self, sheet, row_number, width, read_only):
        self._sheet = sheet
        self.row_number = row_number
        self._width = width
        self._is_read_only = read_only

    def __getitem__(self, slice_obj):
        if isinstance(slice_obj, slice):
            start = slice_obj.start
            stop = slice_obj.stop
            step = slice_obj.step
            if step is None:
                step = 1
            if start is None:
                start = 1
            if stop is None:
                stop = self._width + 1

            return_value = [
                self._sheet.cell(
                    row=self.row_number, column=column + 1
                ).value  # cell() uses 1-based indexes
                for column in range(start, stop, step)
            ]
            if self._is_read_only:
                return strip_whitespace(copy.deepcopy(return_value))
            else:
                return strip_whitespace(return_value)
        else:  # if only retrieving one element
            column = slice_obj + 1  # cell() uses 1-based indexes
            return_value = self._sheet.cell(row=self.row_number, column=column).value
            if self._is_read_only:
                return strip_whitespace(copy.deepcopy(return_value))
            else:
                return strip_whitespace(return_value)

    def __setitem__(self, slice_obj, value):
        if self._is_read_only:
            raise PermissionError(
                f"{full_class_name(self)}.__setitem__: cannot mutate cells of a read only source file object."
            )
        if isinstance(slice_obj, slice):
            start = slice_obj.start
            stop = slice_obj.stop
            step = slice_obj.step
            if step is None:
                step = 1
            if start is None:
                start = 1
            if stop is None:
                stop = self._width + 1
            for col, val in zip(range(start, stop, step), value):
                self._sheet.cell(row=self.row_number, column=col).value = val
        else:
            column = slice_obj + 1
            self._sheet.cell(row=self.row_number, column=column).value = value

    def __iter__(self):
        if self._is_read_only:
            yield from [
                strip_whitespace(copy.deepcopy(cell.value))
                for cell in self._sheet[self.row_number][: self._width + 1]
            ]
        else:
            yield from self._sheet[self.row_number][: self._width + 1]

    def __len__(self):
        return self._width

    def __repr__(self):
        return f"<SourceFileRow Object: {[cell.value for cell in self._sheet[self.row_number][: self._width + 1]]}>"


class SourceFile:
    def __init__(
        self,
        source_file_id,
        attachment_id,
        content,
        max_time_series,
        reporting_year,
        created_by=None,
        read_only=False,
        source_file_name=None,
        extension=None,
        layer_id=None
    ):
        self._source_file_id = source_file_id
        self._attachment_id = attachment_id
        self._content = content
        self._max_time_series = max_time_series
        self._created_by = created_by
        self._is_read_only = read_only
        self.source_file_name = source_file_name
        self._reporting_year = reporting_year
        self._template = source_file_constants.TEMPLATE(reporting_year)

        self._width = (
            source_file_constants.INFO[self._template]['NUM_EMISSION_KEY_COLUMNS']
            + self._max_time_series
            - qc_constants.EARLIEST_REPORTING_YEAR
        )
        self._workbook = None
        self._sheet = None
        self._temp_file = None
        self._is_open = False
        self._source_name_id = None
        self._layer_id = layer_id
        self._extension = extension

    def get_source_file_id(self):
        return self._source_file_id

    def get_attachment_id(self):
        return self._attachment_id

    def get_source_file_name(self):
        return self.source_file_name
    
    def get_template(self):
        return self._template

    def get_max_time_series(self):
        return self._max_time_series

    def get_created_by(self):
        return self._created_by

    def is_read_only(self):
        return self._is_read_only

    def get_source_name_id(self):
        return self._source_name_id
    
    def set_source_name_id(self, id: int):
        self._source_name_id = id

    def get_layer_id(self):
        return self._layer_id
    
    def set_layer_id(self, id: int):
        self._layer_id = id

    def get_reporting_year(self):
        return self._reporting_year
    
    def set_reporting_year(self, year: int):
        self._reporting_year = year

    def get_extension(self):
        return self._extension

    def get_sheet_name_if_exists(self, sheet_name): 
        if self._workbook is None:
            raise PermissionError(
                f"{full_class_name(self)}.get_sheet_name_if_exists: a source file must be opened before its contents can be accessed."
            )
        sheets = self._workbook.sheetnames
        if sheet_name in sheets:
            return sheet_name
        for name in sheets: # check after for case insensitive if doesn't exist
            if sheet_name.lower() == name.lower():
                return name
            
        return None

    def set_active_sheet(self, sheet_name):
        if not self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.set_active_sheet: a source file must be opened before its contents can be accessed."
            )
        if self._sheet.title == sheet_name: # no need to swap
            return
        sheets = self._workbook.sheetnames
        if len(sheets) < 2:
            raise ValueError("Workbook doesn't have enough sheets to change active sheet")
        
        found_sheet_name = self.get_sheet_name_if_exists(sheet_name)

        if found_sheet_name is None:
            raise ValueError(f"Workbook sheet {sheet_name} does not exist to change to")
        self._sheet = self._workbook[found_sheet_name]


    def delete_rows(self, first_row, row_count):
        if not self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.delete_rows: a source file must be opened before its contents can be accessed."
            )
        self._sheet.delete_rows(first_row, row_count)

    def open(self, sheet_name=None):
        if self._is_open:
            helpers.tprint(
                f"Warning: There was an attempt to open a source file that is already open: {self}"
            )
            return
        else:
            self.__enter__(sheet_name)

    def save(self):
        if not self._is_open:
            helpers.tprint(f"Warning: There was an attempt to save a closed source file: {self}")
        self._workbook.save(self._temp_file.name)

    def close(self):
        if not self._is_open:
            helpers.tprint(
                f"Warning: There was an attempt to close a source file that is already closed: {self}"
            )
            return
        else:
            self.__exit__("", "", "")

    # takes csv data and formats it like an excel file for openpyxl
    def prepare_csv(self):
        self._is_open = True
        self._workbook = Workbook()
        self._sheet = self._workbook.active
        self._sheet.title = "InvDB"

        # csv to excel
        file_like_object = io.BytesIO(self._content)
        text_stream = io.TextIOWrapper(file_like_object, encoding='utf-8')
        reader = csv.reader(text_stream)
        gwp_col_pos = source_file_constants.POSITIONS[self.get_template()]['GWP_COL_POS']
        y1990_col_pos = source_file_constants.POSITIONS[self.get_template()]['Y1990_COL_POS']
        for row in reader:
            # convert csv scientific notation strings to floats
            converted_row = []
            for idx, item in enumerate(row):                          
                if idx == gwp_col_pos or idx >= y1990_col_pos:
                    try:
                        converted_item = float(item)
                    except ValueError:
                        converted_item = 0
                else:
                    # no scientific notation outside year columns
                    if item == "NULL" or item == "None" or item == "":
                        converted_item = None
                    else:
                        converted_item = item
                converted_row.append(converted_item)

            self._sheet.append(converted_row)

        self.save()

    # takes json data and formats it like an excel file for openpyxl
    def prepare_json(self):
        self._is_open = True
        self._workbook = Workbook()
        self._sheet = self._workbook.active
        self._sheet.title = "InvDB"

        # json to excel
        json_bytes = io.BytesIO(self._content)
        json_data = json.load(json_bytes)

        format_map = source_file_constants.MAPPINGS.copy()
        for year in range(1990, self.get_max_time_series() + 1):
            format_map[str(year)] = "year"

        for row in json_data:
            # convert csv scientific notation strings to floats
            converted_row = []

            for key in format_map.keys():
                if key in row:
                    item = row[key]
                    if item == "NULL" or item == "None" or item == "":
                        item = None
                    if format_map[key] == "year": # convert year cells to numerics
                        try:
                            item = float(row[key])
                        except ValueError:
                            item = 0
                    converted_row.append(item)

            self._sheet.append(converted_row)

        self.save()

    def get_cell(self, row_num, col_num): 
        return self._sheet[row_num][col_num]

    def __iter__(self):
        if not self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.__iter__: a source file must be opened before it can be iterated."
            )
        row_num = 1
        for row in self._sheet.iter_rows(values_only=True, max_row=self._sheet.max_row):
            if row_num < source_file_constants.INFO[self.get_template()]['FIRST_DATA_ROW']:
                row_num += 1
                continue
            # template 3 files can have blank rows in them, so only stop at blank lines in template 2
            if self._template == 2:
                # stop at the first empty sector value
                if row[source_file_constants.POSITIONS[self.get_template()]['SECTOR_COL_POS']] in [None, "", "NULL"]:
                    return
            if self._template == 3:
                # skip rows with empty sector value
                if row[source_file_constants.POSITIONS[self.get_template()]['SECTOR_COL_POS']] in [None, "", "NULL"]:
                    row_num += 1
                    continue
            # return the next row
            yield SourceFileRow(self._sheet, row_num, self._width, self._is_read_only)
            row_num += 1

    def __getitem__(self, row):
        """retrieves a row from the sheet (note: the row indexes are 1-based)"""
        if not self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.__getitem__: a source file must be opened before its contents can be accessed."
            )
        if not isinstance(row, int):
            raise TypeError(
                f"{full_class_name(self)}.__getitem__: a source file indexes must be integers"
            )
        return SourceFileRow(self._sheet, row, self._width, self._is_read_only)

    def __setitem__(self, _):
        """retrieves a row from the sheet (note: the row indexes are 1-based)"""
        raise AttributeError(
            f"{full_class_name(self)}.__setitem__: row-wise mutations in source files are not supported. Use a second index to mutate slices or individual cells of the row."
        )

    def __enter__(self, sheet_name=None):
        if self._is_open:
            raise PermissionError(
                f"{full_class_name(self)}.__enter__: this file is already open."
            )
        try:
            if self._temp_file is None:
                self._temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=(self._extension or ".xlsx"))
                self._temp_file.write(io.BytesIO(self._content).getbuffer())
                self._temp_file.close()

            if self.get_extension() == ".csv":
                self.prepare_csv()
            elif self.get_extension() == ".json":
                self.prepare_json()
            else:
                self._workbook = load_workbook(self._temp_file.name, data_only=True)
                sheet_name = sheet_name or source_file_constants.SOURCE_FILE_DATA_SHEET_NAME
                found_sheet_name = self.get_sheet_name_if_exists(sheet_name)
                if found_sheet_name is None:
                    raise ValueError(f"Worksheet {sheet_name} does not exist")
                self._sheet = self._workbook[found_sheet_name]
                self._is_open = True                

        except Exception as e:
            helpers.tprint(f"Failed to open workbook: {e}")
            raise

        return self

    def __exit__(self, exc_type, exc_value, traceback):
        try:
            if self._workbook is not None:
                self._workbook.save(self._temp_file.name)
                self._workbook.close()
                self._is_open = False
        except Exception as e:
            helpers.tprint(f"Failed to save/close workbook: {e}")
            raise

        self._workbook = None
        self._sheet = None

        # Remove the temporary file
        # if self._temp_file is not None and os.path.exists(self._temp_file.name):
        #     os.remove(self._temp_file.name)

    def __repr__(self):
        return f"""<SourceFile Object: source file ID: {self._source_file_id}, attachment ID: {self._attachment_id}>"""
